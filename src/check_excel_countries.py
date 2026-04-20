#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
import sys

def check_excel_countries():
    """检查Excel模板中的国家支持"""
    excel_path = 'outputs/成本控制/全球分身规模化成本测算表.xlsx'
    
    try:
        wb = openpyxl.load_workbook(excel_path)
        sheet_names = wb.sheetnames
        print(f"工作表列表: {sheet_names}")
        
        # 检查是否包含国家相关的工作表
        expected_countries = ['US', 'DE', 'SG', 'JP', 'CN']
        found_countries = []
        
        for sheet in sheet_names:
            for country in expected_countries:
                if country in sheet.upper():
                    found_countries.append(country)
        
        print(f"找到的国家: {found_countries}")
        print(f"缺失的国家: {list(set(expected_countries) - set(found_countries))}")
        
        # 检查参数配置表
        if '参数配置表' in sheet_names:
            param_sheet = wb['参数配置表']
            print(f"\n参数配置表检查:")
            # 读取前几行
            for row in param_sheet.iter_rows(min_row=1, max_row=10, values_only=True):
                print(row)
        
        wb.close()
        return True
    except Exception as e:
        print(f"❌ 检查失败: {e}")
        return False

if __name__ == '__main__':
    check_excel_countries()