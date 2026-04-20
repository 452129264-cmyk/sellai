#!/usr/bin/env python3
"""
创建全球成本测算Excel模板
生成包含参数配置表、成本明细表、优化建议表的工作簿
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

def create_global_cost_template(output_path: str = "outputs/成本控制/全球分身规模化成本测算表.xlsx"):
    """
    创建全球成本测算Excel模板
    
    Args:
        output_path: 输出文件路径
    """
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    
    # 移除默认工作表
    if "Sheet" in wb.sheetnames:
        ws = wb["Sheet"]
        wb.remove(ws)
    
    # ====================== 1. 参数配置表 ======================
    ws_params = wb.create_sheet(title="参数配置表")
    
    # 标题
    ws_params.merge_cells('A1:F1')
    title_cell = ws_params['A1']
    title_cell.value = "全球分身规模化成本测算 - 参数配置表"
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 副标题
    ws_params['A3'] = "请填写以下参数开始测算"
    ws_params['A3'].font = Font(size=12, bold=True)
    
    # 参数表格
    headers = ["参数", "值", "单位", "说明", "可选值", "默认值"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_params.cell(row=5, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 参数数据
    parameters = [
        ["目标分身数量", 10, "个", "需要运行的AI分身总数", "1-1000", 10],
        ["平均消息频率", 100, "条/天/分身", "每个分身每天平均处理的消息数", "10-1000", 100],
        ["运营国家", "美国,德国,新加坡", "-", "业务运营的国家（多选）", "北美,欧洲,东南亚,中东,拉美,非洲", "美国,德国,新加坡"],
        ["API调用成本单价", 0.01, "美元/次", "外部API调用平均成本", "0.001-0.1", 0.01],
        ["Token单价", 0.000002, "美元/token", "大模型API的token成本", "0.000001-0.000005", 0.000002],
        ["工作流执行单价", 0.0001, "美元/次", "每次工作流执行的成本", "0.00005-0.0005", 0.0001],
        ["存储单价", 0.000023, "美元/GB/月", "数据存储的月度成本", "0.00001-0.0001", 0.000023],
        ["平均物流成本", 5.0, "美元/订单", "每个订单的平均物流成本", "1.0-50.0", 5.0],
        ["平均关税税率", 10.0, "%", "进口关税税率", "0-50", 10.0],
        ["本地运营费用", 500.0, "美元/月/国家", "每个国家的本地运营费用", "100-5000", 500.0],
        ["运输天数", 7, "天", "平均运输时间", "1-30", 7],
        ["运营月份", 12, "个月", "测算的时间周期", "1-36", 12],
    ]
    
    for row_idx, param in enumerate(parameters, start=6):
        for col_idx, value in enumerate(param, start=1):
            cell = ws_params.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 2:  # 值列，需要用户填写
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    
    # 设置列宽
    for col_idx in range(1, 7):
        ws_params.column_dimensions[get_column_letter(col_idx)].width = 20
    
    # ====================== 2. 成本明细表 ======================
    ws_details = wb.create_sheet(title="成本明细表")
    
    # 标题
    ws_details.merge_cells('A1:H1')
    title_cell = ws_details['A1']
    title_cell.value = "成本明细表 - 按国家地区细分"
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 副标题
    ws_details['A3'] = "基于参数配置表自动计算，请勿手动修改"
    ws_details['A3'].font = Font(size=10, italic=True, color="666666")
    
    # 成本明细表头
    detail_headers = ["国家", "分身数量", "API调用成本", "Token成本", "工作流执行成本", 
                      "存储成本", "物流成本", "关税成本", "本地运营费用", "总成本(美元)"]
    
    for col_idx, header in enumerate(detail_headers, start=1):
        cell = ws_details.cell(row=5, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 示例数据（带有公式）
    # 第6行：美国
    ws_details['A6'] = "美国"
    ws_details['B6'] = '=ROUND(参数配置表!$B$6*0.4,0)'  # 分身数量分配
    ws_details['C6'] = '=$B6*参数配置表!$B$8*30*参数配置表!$B$12'  # API成本
    ws_details['D6'] = '=$B6*参数配置表!$B$7*1000*30*参数配置表!$B$12'  # Token成本
    ws_details['E6'] = '=$B6*参数配置表!$B$9*50*参数配置表!$B$12'  # 工作流成本
    ws_details['F6'] = '=$B6*参数配置表!$B$10*2*参数配置表!$B$12'  # 存储成本
    ws_details['G6'] = '=$B6*参数配置表!$B$13*参数配置表!$B$11'  # 物流成本
    ws_details['H6'] = '=$G6*参数配置表!$B$14/100'  # 关税成本
    ws_details['I6'] = '=参数配置表!$B$15'  # 本地运营费用
    ws_details['J6'] = '=SUM(C6:I6)'  # 总成本
    
    # 第7行：德国
    ws_details['A7'] = "德国"
    ws_details['B7'] = '=ROUND(参数配置表!$B$6*0.35,0)'
    for col in range(3, 10):
        cell = ws_details.cell(row=7, column=col)
        cell.value = f'={get_column_letter(col)}6*0.9'  # 德国成本为美国的90%
    
    # 第8行：新加坡
    ws_details['A8'] = "新加坡"
    ws_details['B8'] = '=ROUND(参数配置表!$B$6*0.25,0)'
    for col in range(3, 10):
        cell = ws_details.cell(row=8, column=col)
        cell.value = f'={get_column_letter(col)}6*0.8'  # 新加坡成本为美国的80%
    
    # 总计行
    ws_details['A9'] = "总计"
    ws_details['B9'] = '=SUM(B6:B8)'
    for col in range(3, 11):
        cell = ws_details.cell(row=9, column=col)
        cell.value = f'=SUM({get_column_letter(col)}6:{get_column_letter(col)}8)'
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # 设置列宽
    for col_idx in range(1, 11):
        ws_details.column_dimensions[get_column_letter(col_idx)].width = 15
    
    # ====================== 3. 优化建议表 ======================
    ws_optimization = wb.create_sheet(title="优化建议表")
    
    # 标题
    ws_optimization.merge_cells('A1:D1')
    title_cell = ws_optimization['A1']
    title_cell.value = "成本优化建议"
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 副标题
    ws_optimization['A3'] = "基于当前测算结果的优化建议"
    ws_optimization['A3'].font = Font(size=12, bold=True)
    
    # 优化建议表头
    opt_headers = ["优化方向", "具体措施", "预计节省成本", "实施难度"]
    for col_idx, header in enumerate(opt_headers, start=1):
        cell = ws_optimization.cell(row=5, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 示例优化建议
    optimizations = [
        ["调整分身分布", "将更多分身部署到低成本区域（如新加坡）", '=成本明细表!J9*0.15', "低"],
        ["优化消息频率", "将平均消息频率降低20%", '=成本明细表!J9*0.10', "中"],
        ["物流整合", "使用集中仓储和批量运输", '=成本明细表!J9*0.08', "高"],
        ["关税优化", "利用自由贸易协定降低税率", '=成本明细表!J9*0.05', "中"],
        ["本地运营共享", "多个国家共享运营团队", '=成本明细表!J9*0.07', "中"],
    ]
    
    for row_idx, opt in enumerate(optimizations, start=6):
        for col_idx, value in enumerate(opt, start=1):
            cell = ws_optimization.cell(row=row_idx, column=col_idx, value=value)
    
    # 总结行
    ws_optimization['A12'] = "总优化潜力"
    ws_optimization['C12'] = '=SUM(C6:C10)'
    ws_optimization['C12'].font = Font(bold=True, color="FF0000")
    
    # 设置列宽
    for col_idx in range(1, 5):
        ws_optimization.column_dimensions[get_column_letter(col_idx)].width = 25
    
    # ====================== 保存工作簿 ======================
    # 确保输出目录存在
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    # 保存文件
    wb.save(output_path)
    
    print(f"全球成本测算Excel模板已创建：{output_path}")
    print("\n工作表说明：")
    print("1. 参数配置表：填写运营参数，黄色单元格为可输入区域")
    print("2. 成本明细表：自动计算各国家成本明细，带有公式")
    print("3. 优化建议表：基于计算结果生成优化建议")
    print("\n使用方法：")
    print("1. 在'参数配置表'中填写运营参数")
    print("2. 查看'成本明细表'获取详细成本分析")
    print("3. 参考'优化建议表'获取优化方向")

if __name__ == "__main__":
    # 生成模板
    create_global_cost_template()