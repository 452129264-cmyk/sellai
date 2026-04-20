#!/usr/bin/env python3
import openpyxl
import sys
import os

def check_excel_structure(filepath):
    """检查Excel文件结构"""
    if not os.path.exists(filepath):
        print(f"文件不存在: {filepath}")
        return
    
    try:
        wb = openpyxl.load_workbook(filepath, data_only=False)
        sheet_names = wb.sheetnames
        print(f"工作表名称: {sheet_names}")
        
        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            print(f"\n工作表: {sheet_name}")
            print(f"  最大行: {ws.max_row}, 最大列: {ws.max_column}")
            
            # 读取前几行内容
            for row in ws.iter_rows(min_row=1, max_row=min(5, ws.max_row), values_only=True):
                print(f"  {row}")
        
        wb.close()
    except Exception as e:
        print(f"读取Excel文件时出错: {e}")

if __name__ == "__main__":
    excel_path = "outputs/成本控制/全球分身规模化成本测算表.xlsx"
    check_excel_structure(excel_path)