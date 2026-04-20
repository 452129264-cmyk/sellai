#!/usr/bin/env python3
import openpyxl
import os
from copy import copy

def update_excel_template(filepath):
    """更新Excel模板，添加日本和中国配置"""
    if not os.path.exists(filepath):
        print(f"文件不存在: {filepath}")
        return False
    
    try:
        wb = openpyxl.load_workbook(filepath, data_only=False)
        
        # 1. 更新参数配置表中的运营国家列表
        ws_param = wb['参数配置表']
        # B8单元格是运营国家
        ws_param['B8'] = '美国,德国,新加坡,日本,中国'
        print(f"更新参数配置表B8为: {ws_param['B8'].value}")
        
        # 2. 更新成本明细表，添加日本和中国行
        ws_cost = wb['成本明细表']
        
        # 当前结构：第6行美国，第7行德国，第8行新加坡，第9行总计
        # 我们需要在第8行后插入两行，然后将总计行下移
        # 但为了简化，我们直接修改第8行为新加坡，第9行为日本，第10行为中国，第11行为总计
        # 首先，我们需要将现有的第8行（新加坡）移动到第9行，第9行（总计）移动到第11行
        # 插入两行在第9行
        
        ws_cost.insert_rows(9, amount=2)  # 在第9行插入两行，原来的第9行（总计）变成第11行
        
        # 现在行号：
        # 6: 美国
        # 7: 德国
        # 8: 新加坡（原来的第8行，现在还在第8行）
        # 9: 新插入的空行（给日本）
        # 10: 新插入的空行（给中国）
        # 11: 总计（原来的第9行）
        
        # 复制新加坡行的格式和公式到日本和中国行
        # 首先获取新加坡行（第8行）的单元格
        source_row = 8  # 新加坡
        
        # 目标行：日本（第9行），中国（第10行）
        for target_row in [9, 10]:
            for col in range(1, ws_cost.max_column + 1):
                source_cell = ws_cost.cell(row=source_row, column=col)
                target_cell = ws_cost.cell(row=target_row, column=col)
                
                # 复制值（公式）
                if source_cell.value is not None:
                    # 替换公式中的行引用
                    if isinstance(source_cell.value, str) and source_cell.value.startswith('='):
                        # 简单的行引用调整：将原公式中的行号6,7,8替换为目标行号？
                        # 实际上，我们需要创建新的公式，使用适当的比例
                        pass
                    else:
                        target_cell.value = source_cell.value
                
                # 复制样式
                if source_cell.has_style:
                    target_cell.font = copy(source_cell.font)
                    target_cell.border = copy(source_cell.border)
                    target_cell.fill = copy(source_cell.fill)
                    target_cell.number_format = source_cell.number_format
                    target_cell.protection = copy(source_cell.protection)
                    target_cell.alignment = copy(source_cell.alignment)
        
        # 设置日本和中国的具体内容
        # 日本行（第9行）
        ws_cost.cell(row=9, column=1, value='日本')
        # 分身数量公式：假设分配比例为15%，四舍五入
        ws_cost.cell(row=9, column=2, value='=ROUND(参数配置表!$B$6*0.15,0)')
        # 其他成本公式：参考新加坡行的公式，但使用不同的调整因子
        # 新加坡行的公式引用C6、D6等，即美国行的成本。对于日本，我们使用不同的比例因子
        # 新加坡的API调用成本公式：=C6*0.8（即美国成本的80%）
        # 对于日本，我们可以设为美国成本的70%？或者独立计算。
        # 但当前模板中，德国和新加坡都引用美国行的成本乘以一个系数。
        # 这意味着我们需要保持这种模式：日本引用美国成本乘以一个系数。
        # 设置日本行的公式（列C到I）
        ws_cost.cell(row=9, column=3, value='=C6*0.7')   # API调用成本：美国成本的70%
        ws_cost.cell(row=9, column=4, value='=D6*0.7')   # Token成本
        ws_cost.cell(row=9, column=5, value='=E6*0.7')   # 工作流执行成本
        ws_cost.cell(row=9, column=6, value='=F6*0.7')   # 存储成本
        ws_cost.cell(row=9, column=7, value='=G6*0.7')   # 物流成本
        ws_cost.cell(row=9, column=8, value='=H6*0.7')   # 关税成本
        ws_cost.cell(row=9, column=9, value='=I6*0.7')   # 本地运营费用
        # 总成本列J公式：=SUM(C9:I9)
        ws_cost.cell(row=9, column=10, value='=SUM(C9:I9)')
        
        # 中国行（第10行）
        ws_cost.cell(row=10, column=1, value='中国')
        ws_cost.cell(row=10, column=2, value='=ROUND(参数配置表!$B$6*0.1,0)')
        ws_cost.cell(row=10, column=3, value='=C6*0.6')   # API调用成本：美国成本的60%
        ws_cost.cell(row=10, column=4, value='=D6*0.6')   # Token成本
        ws_cost.cell(row=10, column=5, value='=E6*0.6')   # 工作流执行成本
        ws_cost.cell(row=10, column=6, value='=F6*0.6')   # 存储成本
        ws_cost.cell(row=10, column=7, value='=G6*0.6')   # 物流成本
        ws_cost.cell(row=10, column=8, value='=H6*0.6')   # 关税成本
        ws_cost.cell(row=10, column=9, value='=I6*0.6')   # 本地运营费用
        ws_cost.cell(row=10, column=10, value='=SUM(C10:I10)')
        
        # 更新新加坡行的比例（原为25%，现在可能保持不变，但总计需要调整）
        # 新加坡行的分身数量公式已经是 =ROUND(参数配置表!$B$6*0.25,0)，保持不变
        
        # 更新美国行和德国行的比例？保持原样：美国40%，德国35%
        # 但总比例现在为：40+35+25+15+10 = 125%，超过100%。需要调整比例。
        # 我们需要重新分配比例，使总和为100%。
        # 更合理的方法是：美国30%，德国25%，新加坡20%，日本15%，中国10%（总和100%）
        # 或者保持原有三国比例不变，将日本和中国作为额外附加。
        # 但成本明细表中的分身数量公式基于目标分身总数乘以比例。
        # 如果比例总和超过100%，会导致分身数量超过目标总数。
        # 所以必须调整比例。
        
        # 决定采用新比例：
        # 美国30%，德国25%，新加坡20%，日本15%，中国10%
        
        # 更新美国行（第6行）分身数量公式
        ws_cost.cell(row=6, column=2, value='=ROUND(参数配置表!$B$6*0.3,0)')
        # 更新德国行（第7行）分身数量公式
        ws_cost.cell(row=7, column=2, value='=ROUND(参数配置表!$B$6*0.25,0)')
        # 更新新加坡行（第8行）分身数量公式
        ws_cost.cell(row=8, column=2, value='=ROUND(参数配置表!$B$6*0.2,0)')
        
        # 更新德国和新加坡的其他成本公式系数（它们引用美国成本）
        # 德国行：原来C6*0.9，现在保持？但美国成本可能变化（因为比例变化），但公式本身不变。
        # 德国行的API成本公式：=C6*0.9（正确）
        # 新加坡行：=C6*0.8（正确）
        # 日本行：=C6*0.7
        # 中国行：=C6*0.6
        
        # 更新总计行（现在在第11行）的公式
        # B列：分身数量总计
        ws_cost.cell(row=11, column=2, value='=SUM(B6:B10)')
        # C列到I列：各成本项总计
        for col in range(3, 11):  # C到J列
            col_letter = openpyxl.utils.get_column_letter(col)
            ws_cost.cell(row=11, column=col, value=f'=SUM({col_letter}6:{col_letter}10)')
        
        # 3. 更新优化建议表（可选，暂无更新）
        
        # 保存文件
        backup_path = filepath.replace('.xlsx', '_backup_20260403.xlsx')
        wb.save(backup_path)
        print(f"已创建备份: {backup_path}")
        
        wb.save(filepath)
        print(f"已更新文件: {filepath}")
        
        return True
        
    except Exception as e:
        print(f"更新Excel模板时出错: {e}")
        return False

if __name__ == "__main__":
    success = update_excel_template("outputs/成本控制/全球分身规模化成本测算表.xlsx")
    if success:
        print("Excel模板更新成功")
    else:
        print("Excel模板更新失败")
        sys.exit(1)