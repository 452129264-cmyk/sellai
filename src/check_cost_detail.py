#!/usr/bin/env python3
import openpyxl
import os

def dump_sheet(filepath, sheet_name):
    """转储工作表内容"""
    if not os.path.exists(filepath):
        print(f"文件不存在: {filepath}")
        return
    
    try:
        wb = openpyxl.load_workbook(filepath, data_only=False)
        ws = wb[sheet_name]
        
        print(f"\n工作表: {sheet_name} (共{ws.max_row}行, {ws.max_column}列)")
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            # 只打印非空行
            if any(cell is not None for cell in row):
                print(row)
        
        wb.close()
    except Exception as e:
        print(f"出错: {e}")

if __name__ == "__main__":
    filepath = "outputs/成本控制/全球分身规模化成本测算表.xlsx"
    dump_sheet(filepath, '成本明细表')
    dump_sheet(filepath, '参数配置表')