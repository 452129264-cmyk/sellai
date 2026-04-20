#!/usr/bin/env python3
import openpyxl
import os
from copy import copy

def update_excel_simple(filepath):
    """简单更新Excel模板，添加日本和中国"""
    if not os.path.exists(filepath):
        print(f"文件不存在: {filepath}")
        return False
    
    try:
        wb = openpyxl.load_workbook(filepath, data_only=False)
        
        # 1. 更新参数配置表：运营国家列表
        ws_param = wb['参数配置表']
        # B8单元格是运营国家
        current_value = ws_param['B8'].value
        print(f"当前运营国家: {current_value}")
        # 添加日本和中国（中文名称）
        new_value = '美国,德国,新加坡,日本,中国'
        ws_param['B8'] = new_value
        print(f"更新后运营国家: {new_value}")
        
        # 2. 更新成本明细表
        ws_cost = wb['成本明细表']
        
        # 当前行结构：
        # 第1行: 标题
        # 第2行: 说明
        # 第3行: 列标题
        # 第4行: 空行？
        # 第5行: 空行？
        # 第6行: 美国
        # 第7行: 德国
        # 第8行: 新加坡
        # 第9行: 总计
        
        # 我们需要插入两行（在第8行后插入，所以原来的第8行新加坡变成第9行？）
        # 实际上，我们可以在第8行（新加坡）和第9行（总计）之间插入两行。
        # 但为了不破坏公式引用，我们直接修改行号。
        # 更简单的方法：删除第9行（总计），然后添加日本、中国、新总计行。
        
        # 首先，保存总计行的公式
        total_row = 9
        total_formulas = {}
        for col in range(1, ws_cost.max_column + 1):
            cell = ws_cost.cell(row=total_row, column=col)
            if cell.value is not None:
                total_formulas[col] = cell.value
        
        # 删除总计行（第9行）
        ws_cost.delete_rows(total_row, 1)
        
        # 现在行号：6美国，7德国，8新加坡
        # 添加日本行（第9行）
        ws_cost.insert_rows(9, 1)
        # 添加中国行（第10行）
        ws_cost.insert_rows(10, 1)
        
        # 设置日本行内容（第9行）
        # 国家名称
        ws_cost.cell(row=9, column=1, value='日本')
        # 分身数量公式：分配比例15%
        ws_cost.cell(row=9, column=2, value='=ROUND(参数配置表!$B$6*0.15,0)')
        # API调用成本公式：引用美国行C6，乘以系数0.7
        ws_cost.cell(row=9, column=3, value='=C6*0.7')
        # Token成本
        ws_cost.cell(row=9, column=4, value='=D6*0.7')
        # 工作流执行成本
        ws_cost.cell(row=9, column=5, value='=E6*0.7')
        # 存储成本
        ws_cost.cell(row=9, column=6, value='=F6*0.7')
        # 物流成本
        ws_cost.cell(row=9, column=7, value='=G6*0.7')
        # 关税成本
        ws_cost.cell(row=9, column=8, value='=H6*0.7')
        # 本地运营费用
        ws_cost.cell(row=9, column=9, value='=I6*0.7')
        # 总成本
        ws_cost.cell(row=9, column=10, value='=SUM(C9:I9)')
        
        # 设置中国行内容（第10行）
        ws_cost.cell(row=10, column=1, value='中国')
        ws_cost.cell(row=10, column=2, value='=ROUND(参数配置表!$B$6*0.1,0)')
        ws_cost.cell(row=10, column=3, value='=C6*0.6')
        ws_cost.cell(row=10, column=4, value='=D6*0.6')
        ws_cost.cell(row=10, column=5, value='=E6*0.6')
        ws_cost.cell(row=10, column=6, value='=F6*0.6')
        ws_cost.cell(row=10, column=7, value='=G6*0.6')
        ws_cost.cell(row=10, column=8, value='=H6*0.6')
        ws_cost.cell(row=10, column=9, value='=I6*0.6')
        ws_cost.cell(row=10, column=10, value='=SUM(C10:I10)')
        
        # 调整原有三国的分配比例，使总和为100%
        # 原比例：美国40%，德国35%，新加坡25% → 总和100%
        # 新比例：美国30%，德国25%，新加坡20%，日本15%，中国10% → 总和100%
        # 更新美国行（第6行）分身数量公式
        ws_cost.cell(row=6, column=2, value='=ROUND(参数配置表!$B$6*0.3,0)')
        # 更新德国行（第7行）
        ws_cost.cell(row=7, column=2, value='=ROUND(参数配置表!$B$6*0.25,0)')
        # 更新新加坡行（第8行）
        ws_cost.cell(row=8, column=2, value='=ROUND(参数配置表!$B$6*0.2,0)')
        
        # 添加新总计行（第11行）
        ws_cost.insert_rows(11, 1)
        # 复制之前的总计行公式，但更新范围
        for col, formula in total_formulas.items():
            if formula is None:
                continue
            # 调整公式中的行范围：原来可能是SUM(C6:C8)，现在应该是SUM(C6:C10)
            if isinstance(formula, str) and formula.startswith('=SUM('):
                # 简单替换：将8替换为10
                new_formula = formula.replace('8', '10')
                ws_cost.cell(row=11, column=col, value=new_formula)
            else:
                ws_cost.cell(row=11, column=col, value=formula)
        
        # 设置总计行国家列
        ws_cost.cell(row=11, column=1, value='总计')
        
        # 确保总计行的公式正确
        # B列：分身数量总计
        ws_cost.cell(row=11, column=2, value='=SUM(B6:B10)')
        # C列到J列：各成本项总计
        col_letters = ['C','D','E','F','G','H','I','J']
        for idx, col in enumerate(range(3, 11), 1):
            col_letter = col_letters[idx-1]
            ws_cost.cell(row=11, column=col, value=f'=SUM({col_letter}6:{col_letter}10)')
        
        # 保存文件
        wb.save(filepath)
        print(f"已更新文件: {filepath}")
        
        # 验证更新
        print("\n验证更新后的成本明细表结构:")
        for row in range(1, 13):
            row_vals = [ws_cost.cell(row=row, column=col).value for col in range(1, 11)]
            if any(v is not None for v in row_vals):
                print(f"第{row}行: {row_vals}")
        
        return True
        
    except Exception as e:
        print(f"更新Excel模板时出错: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    update_excel_simple("outputs/成本控制/全球分身规模化成本测算表.xlsx")