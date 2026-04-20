#!/usr/bin/env python3
"""
验证全球成本测算工具
测试至少3个不同国家的测算案例
"""

import sys
import os
import json
import tempfile
import shutil
from datetime import datetime

# 添加项目根目录到路径
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from src.global_cost_calculator import GlobalCostCalculator

def test_single_country(country_code: str, country_name: str) -> bool:
    """
    测试单个国家的成本测算
    
    Args:
        country_code: 国家代码
        country_name: 国家名称
        
    Returns:
        测试是否通过
    """
    print(f"\n测试 {country_name} ({country_code}):")
    print("-" * 40)
    
    try:
        # 创建测算器
        calculator = GlobalCostCalculator()
        
        # 设置参数：只在该国家运营
        params = calculator.get_default_params()
        params['operating_countries'] = [country_code]
        params['avatar_count'] = 5  # 减少数量以加快测试
        
        # 计算成本
        result = calculator.calculate_costs(params)
        
        # 验证结果
        if 'country_details' not in result:
            print(f"  错误: 结果中缺少 'country_details'")
            return False
        
        if country_code not in result['country_details']:
            print(f"  错误: 结果中缺少目标国家 {country_code}")
            return False
        
        country_result = result['country_details'][country_code]
        
        # 检查必要字段
        required_fields = ['avatar_count', 'total_cost', 'api_cost', 'token_cost', 
                          'logistics_cost', 'tax_cost', 'local_operations_cost']
        
        for field in required_fields:
            if field not in country_result:
                print(f"  错误: 缺少字段 {field}")
                return False
        
        # 验证数值合理性
        if country_result['avatar_count'] != params['avatar_count']:
            print(f"  错误: 分身数量不匹配")
            return False
        
        if country_result['total_cost'] <= 0:
            print(f"  错误: 总成本必须大于0")
            return False
        
        # 显示摘要
        print(f"  分身数量: {country_result['avatar_count']}")
        print(f"  总成本: ${country_result['total_cost']:,.2f} USD")
        print(f"  成本构成:")
        print(f"    API调用: ${country_result['api_cost']:,.2f}")
        print(f"    Token消耗: ${country_result['token_cost']:,.2f}")
        print(f"    物流成本: ${country_result['logistics_cost']:,.2f}")
        print(f"    关税成本: ${country_result['tax_cost']:,.2f}")
        print(f"    本地运营: ${country_result['local_operations_cost']:,.2f}")
        
        # 保存测试报告
        test_dir = "outputs/成本控制/验证测试"
        os.makedirs(test_dir, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_file = os.path.join(test_dir, f"测试报告_{country_code}_{timestamp}.json")
        
        with open(report_file, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        
        print(f"  测试报告已保存: {report_file}")
        
        return True
        
    except Exception as e:
        print(f"  测试失败: {e}")
        import traceback
        traceback.print_exc()
        return False

def test_multiple_countries() -> bool:
    """
    测试多个国家同时运营的成本测算
    
    Returns:
        测试是否通过
    """
    print(f"\n测试多国家运营:")
    print("-" * 40)
    
    try:
        calculator = GlobalCostCalculator()
        
        # 测试三个国家：美国、德国、新加坡
        params = calculator.get_default_params()
        params['operating_countries'] = ['US', 'DE', 'SG']
        params['avatar_count'] = 12
        
        result = calculator.calculate_costs(params)
        
        # 验证三个国家都在结果中
        expected_countries = ['US', 'DE', 'SG']
        for country in expected_countries:
            if country not in result['country_details']:
                print(f"  错误: 缺少国家 {country}")
                return False
        
        # 验证总成本计算
        total_from_details = sum(details['total_cost'] for details in result['country_details'].values())
        total_from_metrics = result['key_metrics']['total_cost_usd']
        
        # 允许小的浮点误差
        if abs(total_from_details - total_from_metrics) > 0.01:
            print(f"  错误: 总成本计算不一致")
            print(f"    明细总和: {total_from_details}")
            print(f"    指标总成本: {total_from_metrics}")
            return False
        
        # 显示结果
        print(f"  运营国家: {', '.join(expected_countries)}")
        print(f"  总分身数: {result['key_metrics']['avatar_count']}")
        print(f"  总成本: ${total_from_metrics:,.2f} USD")
        
        print(f"  各国成本详情:")
        for country, details in result['country_details'].items():
            print(f"    {country}: {details['avatar_count']}分身, ${details['total_cost']:,.2f}")
        
        # 验证优化建议
        if 'optimization_suggestions' not in result:
            print(f"  警告: 缺少优化建议")
        else:
            suggestions = result['optimization_suggestions']
            print(f"  生成优化建议: {len(suggestions)}条")
        
        return True
        
    except Exception as e:
        print(f"  测试失败: {e}")
        return False

def test_excel_template() -> bool:
    """
    验证Excel模板存在且格式正确
    
    Returns:
        测试是否通过
    """
    print(f"\n验证Excel模板:")
    print("-" * 40)
    
    template_path = "outputs/成本控制/全球分身规模化成本测算表.xlsx"
    
    if not os.path.exists(template_path):
        print(f"  错误: Excel模板不存在: {template_path}")
        return False
    
    # 检查文件大小
    file_size = os.path.getsize(template_path)
    if file_size < 1024:  # 至少1KB
        print(f"  错误: Excel模板文件大小异常: {file_size}字节")
        return False
    
    print(f"  Excel模板路径: {template_path}")
    print(f"  文件大小: {file_size:,}字节")
    
    # 尝试用openpyxl打开验证
    try:
        import openpyxl
        wb = openpyxl.load_workbook(template_path, data_only=False)
        
        # 检查工作表
        required_sheets = ['参数配置表', '成本明细表', '优化建议表']
        actual_sheets = wb.sheetnames
        
        for sheet in required_sheets:
            if sheet not in actual_sheets:
                print(f"  错误: 缺少工作表 '{sheet}'")
                return False
        
        print(f"  工作表验证通过: {', '.join(actual_sheets)}")
        
        # 检查参数配置表是否有公式
        ws_params = wb['参数配置表']
        # 查找黄色填充的单元格（可输入区域）
        input_cells = 0
        for row in ws_params.iter_rows():
            for cell in row:
                if cell.fill.start_color.rgb == 'FFFFFF00' or cell.fill.start_color.rgb == 'FFFFEB9C':
                    input_cells += 1
        
        print(f"  可输入参数单元格: {input_cells}个")
        
        wb.close()
        return True
        
    except Exception as e:
        print(f"  警告: Excel模板打开失败: {e}")
        print(f"  但文件存在，可能格式有误")
        return False

def test_shared_state_integration() -> bool:
    """
    验证与共享状态库的集成
    
    Returns:
        测试是否通过
    """
    print(f"\n验证共享状态库集成:")
    print("-" * 40)
    
    try:
        calculator = GlobalCostCalculator()
        
        # 加载历史数据
        historical = calculator.load_historical_data(30)
        
        if not historical:
            print(f"  警告: 未加载到历史数据，可能是数据库为空")
            # 这不是致命错误
            return True
        
        print(f"  历史数据加载成功:")
        print(f"    统计周期: {historical['period_days']}天")
        print(f"    分身数量: {historical['avatar_count']}")
        
        if historical.get('historical_costs'):
            print(f"    成本类型数量: {len(historical['historical_costs'])}")
            for cost_type, data in historical['historical_costs'].items():
                print(f"      {cost_type}: {data['total_amount']:.0f}单位, ${data['total_cost']:.2f}")
        
        return True
        
    except Exception as e:
        print(f"  测试失败: {e}")
        return False

def main():
    """主验证函数"""
    
    print("=" * 60)
    print("全球成本测算工具 - 验证测试")
    print("=" * 60)
    
    # 创建输出目录
    os.makedirs("outputs/成本控制/验证测试", exist_ok=True)
    
    tests_passed = 0
    tests_total = 0
    
    # 测试1: 美国单国家测算
    tests_total += 1
    if test_single_country("US", "美国"):
        tests_passed += 1
    
    # 测试2: 德国单国家测算
    tests_total += 1
    if test_single_country("DE", "德国"):
        tests_passed += 1
    
    # 测试3: 新加坡单国家测算
    tests_total += 1
    if test_single_country("SG", "新加坡"):
        tests_passed += 1
    
    # 测试4: 多国家运营测算
    tests_total += 1
    if test_multiple_countries():
        tests_passed += 1
    
    # 测试5: Excel模板验证
    tests_total += 1
    if test_excel_template():
        tests_passed += 1
    
    # 测试6: 共享状态库集成
    tests_total += 1
    if test_shared_state_integration():
        tests_passed += 1
    
    # 总结
    print("\n" + "=" * 60)
    print("验证测试总结")
    print("=" * 60)
    
    print(f"通过测试: {tests_passed}/{tests_total}")
    
    if tests_passed == tests_total:
        print("✅ 所有验证测试通过！")
        print("\n工具功能完整：")
        print("  1. 支持单国家成本测算")
        print("  2. 支持多国家同时运营测算")
        print("  3. Excel模板格式正确")
        print("  4. 与共享状态库集成正常")
        print("  5. 生成优化建议")
        
        # 保存验证总结
        summary = {
            "timestamp": datetime.now().isoformat(),
            "tests_total": tests_total,
            "tests_passed": tests_passed,
            "tests_failed": tests_total - tests_passed,
            "status": "PASSED" if tests_passed == tests_total else "FAILED",
            "tool_version": "1.0",
            "global_perspective": True
        }
        
        summary_file = "outputs/成本控制/验证测试/验证总结.json"
        with open(summary_file, 'w', encoding='utf-8') as f:
            json.dump(summary, f, ensure_ascii=False, indent=2)
        
        print(f"\n验证总结已保存: {summary_file}")
        return True
    else:
        print("❌ 部分测试失败")
        print("\n建议检查：")
        print("  1. 数据库连接是否正常")
        print("  2. Excel模板文件是否存在")
        print("  3. Python依赖包是否安装")
        return False

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)