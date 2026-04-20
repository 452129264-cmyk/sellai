#!/usr/bin/env python3
import openpyxl
import os

def check_formulas(filepath):
    """检查Excel文件中的公式"""
    if not os.path.exists(filepath):
        print(f"文件不存在: {filepath}")
        return
    
    try:
        wb = openpyxl.load_workbook(filepath, data_only=False)
        ws = wb['成本明细表']
        
        print("检查成本明细表中的公式...")
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith('='):
                    print(f"  单元格 {cell.coordinate}: {cell.value}")
        
        # 检查参数配置表中的国家相关字段
        ws_param = wb['参数配置表']
        print("\n参数配置表中的国家相关字段:")
        for row in ws_param.iter_rows(min_row=1, max_row=ws_param.max_row, values_only=False):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and ('国家' in cell.value or 'country' in cell.value.lower()):
                    print(f"  单元格 {cell.coordinate}: {cell.value}")
        
        wb.close()
    except Exception as e:
        print(f"出错: {e}")

if __name__ == "__main__":
    check_formulas("outputs/成本控制/全球分身规模化成本测算表.xlsx")